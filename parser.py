"""
parser.py — Load GraphPad Prism-style Excel sheets into a tidy long-format DataFrame.

Expected sheet layout
---------------------
Row 1  : "Condition" in A1; treatment names in the first column of each group,
          empty (None) in subsequent replicate columns of the same group.
          Example:  | Condition | DMSO |      |      | 5 µM SalB |      |      |
Rows 2+: One row per experimental condition.
          Example:  | No TEV    | 957250 | 1128682 | ... | 10637817 | ... |

Output
------
pd.DataFrame with columns: condition (str), treatment (str), replicate (int), value (float)
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Union

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_workbook(source: Union[str, Path, bytes, io.BytesIO]) -> openpyxl.Workbook:
    """
    Open an xlsx workbook with cached formula values.
    source may be a file path (str/Path) or raw bytes / BytesIO (from Streamlit uploader).
    """
    if isinstance(source, (str, Path)):
        return openpyxl.load_workbook(source, data_only=True)
    if isinstance(source, (bytes, bytearray)):
        return openpyxl.load_workbook(io.BytesIO(source), data_only=True)
    # assume file-like (BytesIO, UploadedFile …)
    return openpyxl.load_workbook(source, data_only=True)


def sheet_names(wb: openpyxl.Workbook) -> list[str]:
    return wb.sheetnames


def parse_sheet(wb: openpyxl.Workbook, sheet: str) -> pd.DataFrame:
    """
    Parse one sheet and return a tidy DataFrame.
    Returns an empty DataFrame if the sheet has no parseable data.
    """
    ws = wb[sheet]
    return _parse_prism_sheet(ws)


# ---------------------------------------------------------------------------
# Core parsing logic
# ---------------------------------------------------------------------------

def _is_numeric(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _parse_prism_sheet(ws) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame(columns=["condition", "treatment", "replicate", "value"])

    header = rows[0]
    data_rows = rows[1:]

    # ------------------------------------------------------------------
    # Step 1: Find the rightmost column index (0-based) that has any
    #         numeric value in the data rows.  This avoids phantom
    #         trailing columns introduced by Excel formatting.
    # ------------------------------------------------------------------
    max_data_col = 0
    for row in data_rows:
        for i, v in enumerate(row):
            if _is_numeric(v):
                max_data_col = max(max_data_col, i)

    if max_data_col == 0:
        # No numeric data found at all
        return pd.DataFrame(columns=["condition", "treatment", "replicate", "value"])

    # ------------------------------------------------------------------
    # Step 2: Extract treatment names and their starting column indices.
    #         Treatment names appear in non-empty header cells after col 0.
    # ------------------------------------------------------------------
    treatment_starts: list[tuple[int, str]] = []
    for i in range(1, max_data_col + 1):
        val = header[i] if i < len(header) else None
        if val is not None and str(val).strip() != "":
            treatment_starts.append((i, str(val).strip()))

    if not treatment_starts:
        return pd.DataFrame(columns=["condition", "treatment", "replicate", "value"])

    # ------------------------------------------------------------------
    # Step 3: Build treatment spans [start_col, end_col) for each group.
    # ------------------------------------------------------------------
    treatment_spans: list[tuple[int, int, str]] = []
    for idx, (col, name) in enumerate(treatment_starts):
        end = treatment_starts[idx + 1][0] if idx + 1 < len(treatment_starts) else max_data_col + 1
        treatment_spans.append((col, end, name))

    # ------------------------------------------------------------------
    # Step 4: Build long-format records.
    # ------------------------------------------------------------------
    records: list[dict] = []
    for row in data_rows:
        cond_raw = row[0] if row else None
        if cond_raw is None or str(cond_raw).strip() == "":
            continue
        condition = str(cond_raw).strip()

        for start_col, end_col, treatment in treatment_spans:
            rep = 1
            for c in range(start_col, end_col):
                val = row[c] if c < len(row) else None
                if not _is_numeric(val):
                    continue
                records.append(
                    {
                        "condition": condition,
                        "treatment": treatment,
                        "replicate": rep,
                        "value": float(val),
                    }
                )
                rep += 1

    if not records:
        return pd.DataFrame(columns=["condition", "treatment", "replicate", "value"])

    df = pd.DataFrame(records)
    # Preserve the original row order of conditions as a categorical
    cond_order = list(dict.fromkeys(df["condition"]))
    df["condition"] = pd.Categorical(df["condition"], categories=cond_order, ordered=True)
    treat_order = list(dict.fromkeys(df["treatment"]))
    df["treatment"] = pd.Categorical(df["treatment"], categories=treat_order, ordered=True)
    return df


# ---------------------------------------------------------------------------
# Convenience: command-line test
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else None
    if not path:
        print("Usage: python parser.py <file.xlsx>")
        sys.exit(1)

    wb = load_workbook(path)
    print("Sheets:", sheet_names(wb))
    for s in sheet_names(wb):
        df = parse_sheet(wb, s)
        print(f"\n=== {s} === ({len(df)} rows)")
        if df.empty:
            print("  (empty / no parseable numeric data)")
        else:
            print(df.groupby(["condition", "treatment"])["value"].describe().to_string())
